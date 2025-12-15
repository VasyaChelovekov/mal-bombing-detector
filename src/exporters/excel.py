"""
Excel Exporter

Export analysis results to Excel format with formatting and charts.
"""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from src.core.analyzer import AnalysisResult
from src.core.models import SeverityLevel, SuspicionLevel
from src.exporters.base import BaseExporter
from src.utils.config import OutputConfig
from src.utils.i18n import I18nManager


class ExcelExporter(BaseExporter):
    """Export analysis results to Excel format."""
    
    # Color schemes for severity levels
    SEVERITY_COLORS = {
        SeverityLevel.EXTREME: "FF0000",   # Red
        SeverityLevel.SEVERE: "FF6600",    # Orange
        SeverityLevel.MODERATE: "FFCC00",  # Yellow
        SeverityLevel.LIGHT: "99CC00",     # Light green
        SeverityLevel.NONE: "00CC00",      # Green
    }
    
    # Color schemes for suspicion levels
    SUSPICION_COLORS = {
        SuspicionLevel.CRITICAL: "FF0000",  # Red
        SuspicionLevel.HIGH: "FF6600",      # Orange
        SuspicionLevel.MEDIUM: "FFCC00",    # Yellow
        SuspicionLevel.LOW: "00CC00",       # Green
    }
    
    # Header style
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Border style
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    def __init__(
        self,
        output_config: OutputConfig,
        i18n: I18nManager | None = None
    ):
        """Initialize Excel exporter."""
        super().__init__(output_config, i18n)
        
    @property
    def file_extension(self) -> str:
        return "xlsx"
    
    @property
    def name(self) -> str:
        return "Excel"
    
    def export(
        self,
        result: AnalysisResult,
        output_path: Path | str | None = None
    ) -> Path:
        """Export single analysis result to Excel."""
        return self.export_multiple([result], output_path)
    
    def export_multiple(
        self,
        results: list[AnalysisResult],
        output_path: Path | str | None = None
    ) -> Path:
        """
        Export multiple analysis results to Excel.
        
        Creates a workbook with:
        - Summary sheet with all anime
        - Individual sheets for each anime (if enabled)
        - Charts sheet
        """
        if output_path:
            path = Path(output_path)
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = self._get_output_path(f"mal_bombing_analysis_{timestamp}")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        self._create_summary_sheet(wb, results)
        
        # Create detailed metrics sheet
        self._create_metrics_sheet(wb, results)
        
        # Create score distribution sheet
        self._create_distribution_sheet(wb, results)
        
        # Create charts sheet
        self._create_charts_sheet(wb, results)
        
        # Save workbook
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        
        return path
    
    def _create_summary_sheet(
        self,
        wb: Workbook,
        results: list[AnalysisResult]
    ) -> Worksheet:
        """Create summary sheet with all anime."""
        ws = wb.create_sheet(self._translate("export.sheets.summary"))
        
        # Aggregate all metrics from all results
        all_metrics = []
        for r in results:
            all_metrics.extend(r.metrics)
        
        # Prepare data
        data = []
        for m in all_metrics:
            data.append({
                "Rank": m.bombing_rank,
                "Title": m.title,
                "MAL ID": m.mal_id,
                "Bombing Score": round(m.bombing_score, 2),
                "Level": m.suspicion_level.value,
                "Ones %": round(m.ones_percentage, 2),
                "Tens %": round(m.tens_percentage, 2),
                "Ones Z-Score": round(m.ones_zscore, 2),
                "Spike Ratio": round(m.spike_ratio, 2),
                "Effect Size": round(m.distribution_effect_size, 3),
            })
        
        df = pd.DataFrame(data)
        
        # Write to worksheet
        self._write_dataframe(ws, df)
        
        # Apply severity colors
        level_col = self._get_column_index(df, "Level")
        self._apply_severity_colors(ws, level_col, all_metrics)
        
        # Adjust column widths
        self._auto_fit_columns(ws)
        
        return ws
    
    def _create_metrics_sheet(
        self,
        wb: Workbook,
        results: list[AnalysisResult]
    ) -> Worksheet:
        """Create detailed metrics sheet."""
        ws = wb.create_sheet(self._translate("export.sheets.metrics"))
        
        # Aggregate all metrics
        all_metrics = []
        for r in results:
            all_metrics.extend(r.metrics)
        
        # Prepare data with all metrics
        data = []
        for m in all_metrics:
            row = {
                "Title": m.title,
                "Bombing Score": round(m.bombing_score, 2),
                "Ones Z-Score": round(m.ones_zscore, 2),
                "Spike Ratio": round(m.spike_ratio, 2),
                "Effect Size": round(m.distribution_effect_size, 3),
                "Entropy Deficit": round(m.entropy_deficit, 3),
                "Bimodality": round(m.bimodality_coefficient, 3),
            }
            
            # Add metric breakdown
            for key, value in m.metric_breakdown.items():
                row[f"Score: {key}"] = round(value, 3)
            
            data.append(row)
        
        df = pd.DataFrame(data)
        self._write_dataframe(ws, df)
        self._auto_fit_columns(ws)
        
        return ws
        
        return ws
    
    def _create_distribution_sheet(
        self,
        wb: Workbook,
        results: list[AnalysisResult]
    ) -> Worksheet:
        """Create score distribution sheet."""
        ws = wb.create_sheet(self._translate("export.sheets.distribution"))
        
        # Headers
        headers = ["Title", "MAL ID"]
        headers.extend([f"Score {i} %" for i in range(1, 11)])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # Aggregate all metrics
        all_metrics = []
        for r in results:
            all_metrics.extend(r.metrics)
        
        # Note: We don't have distribution data in ReviewBombingMetrics
        # Just show ones_percentage and tens_percentage
        for row_idx, m in enumerate(all_metrics, 2):
            ws.cell(row=row_idx, column=1, value=m.title).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=2, value=m.mal_id).border = self.THIN_BORDER
            
            # We only have ones% and tens%
            ws.cell(row=row_idx, column=3, value=round(m.ones_percentage, 2)).border = self.THIN_BORDER
            ws.cell(row=row_idx, column=12, value=round(m.tens_percentage, 2)).border = self.THIN_BORDER
        
        self._auto_fit_columns(ws)
        
        return ws
    
    def _create_charts_sheet(
        self,
        wb: Workbook,
        results: list[AnalysisResult]
    ) -> Worksheet:
        """Create charts sheet."""
        ws = wb.create_sheet(self._translate("export.sheets.charts"))
        
        # Only create chart for top suspicious anime
        suspicious = []
        for r in results:
            for m in r.metrics:
                if m.suspicion_level in [
                    SuspicionLevel.CRITICAL,
                    SuspicionLevel.HIGH,
                    SuspicionLevel.MEDIUM
                ]:
                    suspicious.append(m)
        suspicious = suspicious[:5]
        
        if not suspicious:
            ws.cell(row=1, column=1, value="No suspicious anime found for charts")
            return ws
        
        # Prepare chart data
        ws.cell(row=1, column=1, value="Bombing Score Comparison")
        ws.cell(row=2, column=1, value="Title")
        ws.cell(row=2, column=2, value="Score")
        
        for idx, m in enumerate(suspicious, 3):
            ws.cell(row=idx, column=1, value=m.title[:30])
            ws.cell(row=idx, column=2, value=m.bombing_score)
        
        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Top Suspicious Anime"
        chart.x_axis.title = "Anime"
        chart.y_axis.title = "Bombing Score"
        
        data = Reference(ws, min_col=2, min_row=2, max_row=2 + len(suspicious))
        categories = Reference(ws, min_col=1, min_row=3, max_row=2 + len(suspicious))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.shape = 4
        chart.width = 15
        chart.height = 10
        
        ws.add_chart(chart, "D2")
        
        return ws
    
    def _write_dataframe(self, ws: Worksheet, df: pd.DataFrame) -> None:
        """Write DataFrame to worksheet with formatting."""
        # Headers
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        
        # Data rows
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER
                
                # Number formatting
                if isinstance(value, float):
                    cell.number_format = "#,##0.00"
                elif isinstance(value, int):
                    cell.number_format = "#,##0"
    
    def _get_column_index(self, df: pd.DataFrame, column_name: str) -> int:
        """Get 1-based column index by name."""
        try:
            return list(df.columns).index(column_name) + 1
        except ValueError:
            return -1
    
    def _apply_severity_colors(
        self,
        ws: Worksheet,
        col_idx: int,
        metrics_list: list
    ) -> None:
        """Apply background colors based on suspicion level."""
        if col_idx < 0:
            return
        
        # Map suspicion levels to colors
        level_colors = {
            SuspicionLevel.CRITICAL: "FF0000",
            SuspicionLevel.HIGH: "FF6600",
            SuspicionLevel.MEDIUM: "FFCC00",
            SuspicionLevel.LOW: "00CC00",
        }
            
        for row_idx, m in enumerate(metrics_list, 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            color = level_colors.get(m.suspicion_level, "FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            # Dark text for light backgrounds (LOW level)
            if m.suspicion_level == SuspicionLevel.LOW:
                cell.font = Font(color="000000")
            else:
                cell.font = Font(color="FFFFFF", bold=True)
    
    def _auto_fit_columns(self, ws: Worksheet) -> None:
        """Auto-fit column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except (TypeError, AttributeError):
                    pass
            
            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
